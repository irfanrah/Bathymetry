import matplotlib as mpl
from mpl_toolkits.mplot3d import Axes3D
import matplotlib.pyplot as plt
import time
import numpy as np
import random
from matplotlib import cm
from openpyxl import load_workbook
import math
from scipy import interpolate



kebenaran = True
kebenaranz = True
wb = load_workbook(filename = 'olda.xlsx')
fig = plt.figure()
ax = fig.add_subplot(111, projection='3d')


sheet = wb['Sheet1']

SumbuX = []
SumbuX2 = []
SumbuY = []
SumbuY2 = []
SumbuZ = []

i = 3

while (kebenaran):
	DatX = sheet.cell(row=i, column=11).value
	DatY = sheet.cell(row=i, column=12).value
	DatZ = sheet.cell(row=i, column=5).value
	if DatX == None :
		kebenaran = False
	else :
		SumbuX.append(DatX)
		SumbuY.append(DatY)
		SumbuZ.append(DatZ)
	i = i + 1
	

x1 = math.floor(min(SumbuX))
y1 = math.floor(min(SumbuY))
x2 = math.ceil(max(SumbuX))
y2 = math.ceil(max(SumbuY))


for i in range(len(SumbuZ)):
	
	Xdot = math.ceil(SumbuX[i])
	Ydot = math.ceil(SumbuY[i])
	
	if Xdot < 0 :
		CalcX = Xdot*2*(-1) + Xdot
		if Ydot < 0:
			CalcY = Ydot*2*(-1) + Ydot
		else :
			CalcY = Ydot*2 + Ydot
	else :
		CalcX = Xdot*2 + Xdot
		if Ydot < 0:
			CalcY = Ydot*2*(-1) + Ydot
		else :
			CalcY = Ydot*2 + Ydot
	SumbuX2.append(CalcX)
	SumbuY2.append(CalcY)

arrX = np.asarray(SumbuX2)
arrY = np.asarray(SumbuY2)
arrZ = np.asarray(SumbuZ)




minimalX = min(arrX)
maximalX = max(arrX)
minimalY = min(arrY)
maximalY = max(arrY)


X, Y = np.meshgrid(arrX, arrY)
grid_z0 = interpolate.griddata((arrX,arrY), arrZ, (xm, ym), method='nearest')
surf = ax.plot_surface(X, Y, grid_z0, cmap=cm.coolwarm, linewidth=0, antialiased=False)
plt.show()
