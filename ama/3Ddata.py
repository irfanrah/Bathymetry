from openpyxl import load_workbook
from mpl_toolkits.mplot3d import Axes3D  # noqa: F401 unused import
import matplotlib.pyplot as plt
import numpy as np
import math
from matplotlib import cm
from matplotlib.ticker import LinearLocator, FormatStrFormatter



kebenaran = True
kebenaranz = True
wb = load_workbook(filename = 'olda.xlsx')
fig = plt.figure()
ax = fig.add_subplot(111, projection='3d')


sheet = wb['Sheet1']

SumbuX = []
SumbuX2 = []
SumbuY = []
SumbuY2 = []
SumbuZ = []

i = 3

while (kebenaran):
	DatX = sheet.cell(row=i, column=11).value
	DatY = sheet.cell(row=i, column=12).value
	DatZ = sheet.cell(row=i, column=5).value
	if DatX == None :
		kebenaran = False
	else :
		SumbuX.append(DatX)
		SumbuY.append(DatY)
		SumbuZ.append(DatZ)
	i = i + 1
	

x1 = math.floor(min(SumbuX))
y1 = math.floor(min(SumbuY))
x2 = math.ceil(max(SumbuX))
y2 = math.ceil(max(SumbuY))

X = np.arange(0,400, 1)
Y = np.arange(0, 400, 1)
X, Y = np.meshgrid(X, Y)
Z = np.zeros(shape=(400,400))
print(x2,y2)

for i in range(len(SumbuZ)):
	
	Xdot = math.ceil(SumbuX[i])
	#
	Ydot = math.ceil(SumbuY[i])
	
	if Xdot < 0 :
		CalcX = Xdot*2*(-1) + Xdot
		if Ydot < 0:
			CalcY = Ydot*2*(-1) + Ydot
		else :
			CalcY = Ydot*2 + Ydot
		# Z[ Xdot*2*(-1) + Xdot ,Ydot*2*(-1) + Ydot] = SumbuZ[i]
	else :
		CalcX = Xdot*2 + Xdot
		if Ydot < 0:
			CalcY = Ydot*2*(-1) + Ydot
		else :
			CalcY = Ydot*2 + Ydot
	SumbuX2.append(CalcX)
	SumbuY2.append(CalcY)


# Plot the surface.
#surf = ax.plot_surface(X, Y, Z, cmap=cm.coolwarm,
                       #linewidth=0, antialiased=False)

# Customize the z axis.
#ax.zaxis.set_major_locator(LinearLocator(10))

# Add a color bar which maps values to colors.
#fig.colorbar(surf, shrink=0.5, aspect=5)

ax.scatter(SumbuX2, SumbuY2, SumbuZ)
plt.show()
